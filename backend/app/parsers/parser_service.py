from typing import List, Dict, Any
from pathlib import Path
import docx
from pypdf import PdfReader
from pypdf.errors import PdfReadError
from bs4 import BeautifulSoup
import openpyxl
import logging

from app.models import Artifact

logger = logging.getLogger(__name__)


class ParserService:
    """Service for parsing and chunking different document types."""

    async def parse_and_chunk(self, artifact: Artifact) -> List[Dict[str, Any]]:
        """Parse artifact and return chunks with locators."""
        if artifact.type == "docx":
            return self._parse_docx(artifact)
        elif artifact.type == "pdf":
            return self._parse_pdf(artifact)
        elif artifact.type in ["txt", "md"]:
            return self._parse_text(artifact)
        elif artifact.type == "url":
            return self._parse_text(artifact)
        elif artifact.type == "xlsx":
            return self._parse_xlsx(artifact)
        else:
            raise ValueError(f"Unsupported artifact type: {artifact.type}")

    def _parse_docx(self, artifact: Artifact) -> List[Dict[str, Any]]:
        """Parse DOCX file and chunk by headings."""
        doc = docx.Document(artifact.source_path)
        chunks = []
        current_heading_path = []
        current_section_text = []
        para_index = 0

        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue

            # Check if paragraph is a heading
            if para.style.name.startswith('Heading'):
                # Save previous section if exists
                if current_section_text:
                    chunks.append({
                        "text": "\n".join(current_section_text),
                        "locator": {
                            "type": "docx",
                            "heading_path": current_heading_path.copy(),
                            "para_start": para_index - len(current_section_text),
                            "para_end": para_index
                        }
                    })
                    current_section_text = []

                # Update heading path
                heading_level = int(para.style.name.replace('Heading', '').strip() or '1')
                current_heading_path = current_heading_path[:heading_level-1] + [text]

            current_section_text.append(text)
            para_index += 1

        # Add final section
        if current_section_text:
            chunks.append({
                "text": "\n".join(current_section_text),
                "locator": {
                    "type": "docx",
                    "heading_path": current_heading_path.copy(),
                    "para_start": para_index - len(current_section_text),
                    "para_end": para_index
                }
            })

        return chunks

    def _parse_pdf(self, artifact: Artifact) -> List[Dict[str, Any]]:
        """Parse PDF file and chunk by pages."""
        chunks = []
        
        try:
            reader = PdfReader(artifact.source_path)
            
            for page_num, page in enumerate(reader.pages, start=1):
                try:
                    text = page.extract_text()
                    if not text or not text.strip():
                        continue

                    # Split page into paragraphs (simple heuristic)
                    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]

                    # Create chunks for substantial paragraphs
                    for para_idx, para in enumerate(paragraphs):
                        if len(para) > 50:  # Minimum chunk size
                            chunks.append({
                                "text": para,
                                "locator": {
                                    "type": "pdf",
                                    "page": page_num,
                                    "para_idx": para_idx,
                                    "total_pages": len(reader.pages)
                                }
                            })
                except Exception as e:
                    logger.warning(f"Failed to extract text from page {page_num}: {e}")
                    continue
                    
        except PdfReadError as e:
            logger.error(f"PDF read error for {artifact.source_path}: {e}")
            # Return an error chunk instead of crashing
            chunks.append({
                "text": f"[ERROR: Unable to parse PDF - file may be corrupted or password-protected. Error: {str(e)}]",
                "locator": {
                    "type": "pdf",
                    "page": 0,
                    "para_idx": 0,
                    "error": True
                }
            })
        except Exception as e:
            logger.error(f"Unexpected error parsing PDF {artifact.source_path}: {e}")
            chunks.append({
                "text": f"[ERROR: Unexpected error parsing PDF: {str(e)}]",
                "locator": {
                    "type": "pdf",
                    "page": 0,
                    "para_idx": 0,
                    "error": True
                }
            })
        
        return chunks

    def _parse_text(self, artifact: Artifact) -> List[Dict[str, Any]]:
        """Parse plain text or markdown file."""
        with open(artifact.source_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Split by double newlines (paragraphs)
        paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
        chunks = []

        current_heading = None
        for para_idx, para in enumerate(paragraphs):
            # Detect markdown headings
            if para.startswith('#'):
                current_heading = para.lstrip('#').strip()

            if len(para) > 50:
                chunks.append({
                    "text": para,
                    "locator": {
                        "type": artifact.type,
                        "para_idx": para_idx,
                        "heading": current_heading
                    }
                })

        return chunks

    def _parse_xlsx(self, artifact: Artifact) -> List[Dict[str, Any]]:
        """Parse Excel file and chunk by rows."""
        workbook = openpyxl.load_workbook(artifact.source_path, data_only=True)
        chunks = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]

            # Process each row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Create text representation of row
                row_data = []
                for header, value in zip(headers, row):
                    if value:
                        row_data.append(f"{header}: {value}")

                if row_data:
                    chunks.append({
                        "text": " | ".join(row_data),
                        "locator": {
                            "type": "xlsx",
                            "sheet": sheet_name,
                            "row": row_idx
                        }
                    })

        return chunks
